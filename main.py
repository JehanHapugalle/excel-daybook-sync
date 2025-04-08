import os
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()  # Load environment variables from .env file to manage file paths securely

# Folder path where the coupon files are stored
coupon_directory = os.getenv('COUPON_FOLDER_PATH')
# Path to the Excel file where the rows will be saved
output_file = os.getenv('OUTPUT_FILE_PATH')

def count_red_text_in_november(coupon_directory, output_file):
    total_red_text_count = 0  # Track numer of rows with red text within the current month found across all sheets
    rows_to_append = []  # List storing rows meeting the condition

    # Loop through all files in the folder
    for file_name in os.listdir(coupon_directory):
        # Only process Excel files ending with .xlsx) and contain "coupons" in name
        if file_name.lower().endswith(".xlsx") and "coupons" in file_name.lower():
            try:
                # Open current tutor coupon Excel file within the loop
                wb = load_workbook(os.path.join(coupon_directory, file_name))

                # Loop through all sheets in the current Excel file
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Looks for first instance of "date" string in column B to identify where coupon records start.
                    date_header_row = next((idx for idx, row in enumerate(ws.iter_rows(min_col=2, max_col=2, values_only=True), start=1) 
                                            if row[0] and "date" in str(row[0]).lower()), None)
                    
                    if date_header_row:  # If date header found
                        red_text_count = 0  # Count number of rows with red text found within current month
                        red_text_rows = []  # Store which row numbers have red text within current month

                        # Loop through all rows below date header and check the values
                        for row_idx, row in enumerate(ws.iter_rows(min_row=date_header_row + 1, min_col=1, max_col=ws.max_column, values_only=True), start=date_header_row + 1):
                            cell_value_b = row[1]  # Get value from column B (Date)
                            cell_value_a = row[0]  # Get value from column A (Red text check)
                            
                            # If date is in April 2025
                            if isinstance(cell_value_b, datetime) and cell_value_b.month == 4 and cell_value_b.year == 2025:
                                # Check if cell in column A has red text
                                cell_a = ws.cell(row=row_idx, column=1)  # Access corresponding cell in column A to check text color
                                if cell_a.value and cell_a.font.color and cell_a.font.color.rgb == 'FFFF0000':  # Checks if text colour is Red (FFFF0000)
                                    red_text_count += 1  # Increase count for red text rows
                                    red_text_rows.append(row_idx)  # Save row number
                                    rows_to_append.append(row)  # Save entire row to append later

                        # If any red text rows found, print result
                        if red_text_count > 0:
                            print(f"\n{sheet_name}: {red_text_count} 10% free coupons (rows with red text in column A and date within current month)")
                            print(f"  Row(s): {', '.join(map(str, red_text_rows))}")
                            
                            total_red_text_count += red_text_count  # Add to overall red text row count

            except Exception as e:
                print(f"\nError processing file {file_name}: {e}") # Error message with description if unable to process file. Note: Tutor coupon files must be closed if not will cause error.

    # After processing all files, print total red text row count
    print(f"\nTotal 10% free coupons across all sheets: {total_red_text_count}")

    # Append rows we collected to the output Excel file
    try:
        # Open the output day book Excel file
        output_wb = load_workbook(output_file) # Output filepath variable
        output_ws = output_wb.active  # Access active sheet in the output file

        # Append rows that met the red text condition
        for row in rows_to_append:
            output_ws.append(row)

        # Save changes to the output file
        output_wb.save(output_file)
        print(f"\nRows successfully appended to {output_file}")
    except Exception as e:
        print(f"Error saving to output file: {e}") # Error message with description if unable to save to tutor day book. Note: Day book file must be closed, else error.

# Call function to execute the task
count_red_text_in_november(coupon_directory, output_file)